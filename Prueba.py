# Horarios totales
from openpyxl import Workbook
from openpyxl import load_workbook



ruta_guardado = "C:\\Users\\Juan Martin\\Desktop\\Juan Martin\\Github\\Prueba\\Horarios\\horas.xlsx"
libro = Workbook()
hoja = libro.active
libro2 = load_workbook(ruta_guardado)
hoja2 = libro2["Sheet"]
celda = hoja2["A2"]
valor = celda.value
cerrar = 0
print("Usted tiene registradas: " ) 
print(valor )

b = input("Ponga un punto para sumar 4h  ")
while b == ".":
    if "." == b:
       valor += 4
    b = input("Ponga un punto para sumar 4h  ")

hoja['A1'] = 'Cant. horas'
hoja['A2'] = valor

libro.save(ruta_guardado)

print("")
print("Ya se han guardado en tu excel las horas.")
print("")


while cerrar == 0:
    print("----------------------------")
    cerrar = input("Presione enter para cerrar | \n----------------------------")
   

 


